from pathlib import Path
from typing import List
from datetime import datetime
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from ..models.filtered_domain import FilteredDomain

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Экспорт данных в Excel с форматированием"""
    
    @staticmethod
    def export(
        domains: List[FilteredDomain],
        output_file: str,
        include_spam: bool = False,
        include_excluded: bool = False,
        target_domain: str = None
    ) -> str:
        """
        Экспорт доменов в Excel
        
        Args:
            domains: Список доменов
            output_file: Путь к выходному файлу
            include_spam: Включить спам-домены
            include_excluded: Включить исключенные домены
            target_domain: Целевой домен (для титульной страницы)
            
        Returns:
            Путь к созданному файлу
        """
        if not EXCEL_AVAILABLE:
            raise ImportError(
                "openpyxl не установлен. Установите: pip install openpyxl"
            )
        
        logger.info(f"Экспорт в Excel: {output_file}")
        
        # Фильтруем домены
        filtered = [
            d for d in domains
            if (include_spam or not d.is_spam) and
               (include_excluded or not d.is_excluded) and
               d.is_registered
        ]
        
        # Создаем workbook
        wb = Workbook()
        
        # Лист 1: Сводка
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ExcelExporter._create_summary_sheet(
            ws_summary,
            domains,
            filtered,
            target_domain
        )
        
        # Лист 2: Данные
        ws_data = wb.create_sheet("Domains")
        ExcelExporter._create_data_sheet(ws_data, filtered)
        
        # Сохранение
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)
        
        logger.info(f"✓ Экспортировано {len(filtered)} доменов в {output_path}")
        return str(output_path)
    
    @staticmethod
    def _create_summary_sheet(ws, all_domains, filtered_domains, target_domain):
        """Создание сводного листа"""
        # Заголовок
        ws['A1'] = "Domain Backlink Analysis Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Информация о целевом домене
        row = 3
        if target_domain:
            ws[f'A{row}'] = "Target Domain:"
            ws[f'B{row}'] = target_domain
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        ws[f'A{row}'] = "Report Date:"
        ws[f'B{row}'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        # Статистика
        ws[f'A{row}'] = "Statistics:"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        stats = [
            ("Total Domains Analyzed", len(all_domains)),
            ("Valid Domains", len(filtered_domains)),
            ("Spam Domains", sum(1 for d in all_domains if d.is_spam)),
            ("Excluded Domains", sum(1 for d in all_domains if d.is_excluded)),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Топ домены
        row += 1
        ws[f'A{row}'] = "Top 10 Domains by DR:"
        ws[f'A{row}'].font = Font(size=12, bold=True)
        row += 1
        
        sorted_domains = sorted(
            filtered_domains,
            key=lambda d: d.dr if d.dr else 0,
            reverse=True
        )[:10]
        
        ws[f'A{row}'] = "Domain"
        ws[f'B{row}'] = "DR"
        ws[f'C{row}'] = "Backlinks"
        for cell in [ws[f'A{row}'], ws[f'B{row}'], ws[f'C{row}']]:
            cell.font = Font(bold=True)
        row += 1
        
        for domain in sorted_domains:
            ws[f'A{row}'] = domain.domain
            ws[f'B{row}'] = domain.dr or 'N/A'
            ws[f'C{row}'] = domain.backlink_count
            row += 1
        
        # Автоподбор ширины
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 30
    
    @staticmethod
    def _create_data_sheet(ws, domains):
        """Создание листа с данными"""
        # Заголовки
        headers = [
            'Domain', 'DR', 'UR', 'Backlink Count',
            'Total Backlinks', 'Referring Domains',
            'Organic Traffic', 'Spam Status', 'Excluded'
        ]
        
        # Стили для заголовков
        header_fill = PatternFill(
            start_color="366092",
            end_color="366092",
            fill_type="solid"
        )
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Данные
        for row_idx, domain in enumerate(domains, 2):
            ws.cell(row=row_idx, column=1, value=domain.domain)
            ws.cell(row=row_idx, column=2, value=domain.dr or 'N/A')
            ws.cell(row=row_idx, column=3, value=domain.ur or 'N/A')
            ws.cell(row=row_idx, column=4, value=domain.backlink_count)
            ws.cell(row=row_idx, column=5, value=domain.total_backlinks or 'N/A')
            ws.cell(row=row_idx, column=6, value=domain.referring_domains or 'N/A')
            ws.cell(row=row_idx, column=7, value=domain.organic_traffic or 'N/A')
            ws.cell(row=row_idx, column=8, value='SPAM' if domain.is_spam else 'CLEAN')
            ws.cell(row=row_idx, column=9, value='YES' if domain.is_excluded else 'NO')
            
            # Цветовая кодировка для спама
            if domain.is_spam:
                for col in range(1, 10):
                    ws.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFE6E6",
                        end_color="FFE6E6",
                        fill_type="solid"
                    )
        
        # Автоподбор ширины колонок
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
